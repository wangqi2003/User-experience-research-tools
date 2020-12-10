from snownlp import SnowNLP

import openpyxl
# 打开excel
excel = openpyxl.load_workbook('D:\\#王琦\\方太烟灶-文本分析\\海尔\\haier.xlsx') #华帝\\vatti.xlsx    老板\\robam.xlsx   美的\\midea.xlsx
# 使用指定工作表
sheet = excel.active  # 当前激活的工作表
L=sheet.max_row
print(L)

new_file_positive=openpyxl.Workbook()
ws=new_file_positive.active

ws['A1']='评论内容'
ws['B1']='正向情感评分'
  
new_file_negative=openpyxl.Workbook()
ws2=new_file_negative.active

ws2['A1']='评论内容'
ws2['B1']='负向情感评分'


j1=1
j2=1


for i in range(1,L+1):
    print(i)
    text=sheet.cell(i,1).value

    #""" 调用情感倾向分析 """
    outcome=SnowNLP(text)

    score=outcome.sentiments


    if score>=0.01:#积极
        j1 += 1
        ws.cell(j1,1,text)
        ws.cell(j1,2,score)
       

    elif score<0.01:#消极
        j2 += 1
        ws2.cell(j2,1,text)
        ws2.cell(j2,2,score)
        

new_file_positive.save('D:\\#王琦\\方太烟灶-文本分析\\海尔\\haier_com_hao.xlsx')  #老板\\robam   华帝\\vatti   美的\\midea
new_file_negative.save('D:\\#王琦\\方太烟灶-文本分析\\海尔\\haier_com_cha.xlsx')