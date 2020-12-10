import openpyxl
# 打开excel
excel = openpyxl.load_workbook('D:\\#王琦\\方太烟灶-文本分析\\京东产品评论_清洗_保留中文.xlsx')#全部数据表
# 使用指定工作表
sheet = excel.active  # 当前激活的工作表
L=sheet.max_row
print(L)

#SKU=['29432213661','64607790377'.'50037984843']  华帝（VATTI）  方太（FOTILE）  美的（midea）   海尔（Haier）  百得（BEST）  老板（Robam）

Robam=openpyxl.Workbook()
ws=Robam.active

#SKU64607790377=openpyxl.Workbook()
#ws=new_file_positive.active

#SKU50037984843=openpyxl.Workbook()
#ws=new_file_positive.active

j=1

for i in range(2,L+2):
    if sheet.cell(i,7).value=='海尔（Haier）':
        comment = sheet.cell(i,15).value#评论文本
        #print(comment)
        ws.cell(j,1,comment)
        j += 1
print(j)
Robam.save('D:\\#王琦\\方太烟灶-文本分析\\海尔\\haier.xlsx')
Robam.close()