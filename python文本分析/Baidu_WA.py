
from aip import AipNlp
import time
import chardet
APP_ID = '18681346'
API_KEY = 'WFPdoYeukds3UdZTjDTETtHM'
SECRET_KEY = 'S49YyoGwi7XpTcyAHsGfFcmzQMQuADCt'

client = AipNlp(APP_ID, API_KEY, SECRET_KEY)

import openpyxl
#with open('D:\\#王琦\\方太烟灶-文本分析\\美的\\Midea.xlsx', 'rb') as file:
     #print(chardet.detect(file.read()))
#file.close()

# 打开excel
excel = openpyxl.load_workbook('D:\\#王琦\\方太烟灶-文本分析\\华帝\\vatti.xlsx')  #华帝\\vatti.xlsx    老板\\robam.xlsx   美的\\midea.xlsx
# 使用指定工作表

sheet= excel.active  # 当前激活的工作表
L=sheet.max_row
print(L)

new_file_positive=openpyxl.Workbook()
ws=new_file_positive.active

ws['A1']='评论内容'
ws['B1']='正向情感概率'
ws['C1']='负向情感概率'
  

new_file_negative=openpyxl.Workbook()
ws2=new_file_negative.active

ws2['A1']='评论内容'
ws2['B1']='正向情感概率'
ws2['C1']='负向情感概率'

new_file_nutural=openpyxl.Workbook()
ws3=new_file_nutural.active

ws3['A1']='评论内容'
ws3['B1']='正向情感概率'
ws3['C1']='负向情感概率'

j=1
x=1
k=1
for i in range(1,L+1):
    print(i)
    text=sheet.cell(i,1).value
    time.sleep(0.5)

    #""" 调用情感倾向分析 """
    outcome=client.sentimentClassify(text)
    #print(client.sentimentClassify(text))

    score=outcome['items'][0]['sentiment']
    print(score)
    posi=outcome['items'][0]['positive_prob']
    #print(posi)
    nega=outcome['items'][0]['negative_prob']
    #print(nega)

    
    if score==2:#积极
        j += 1
        ws.cell(j,1,text)
        ws.cell(j,2,posi)
        ws.cell(j,3,nega)

    elif score==0:#消极
        x += 1
        ws2.cell(x,1,text)
        ws2.cell(x,2,posi)
        ws2.cell(x,3,nega)
    else:
        k += 1
        ws3.cell(k,1,text)
        ws3.cell(k,2,posi)
        ws3.cell(k,3,nega)
    
new_file_positive.save('D:\\#王琦\\方太烟灶-文本分析\\华帝\\vatti_com_positive.xlsx')  #老板\\robam   华帝\\vatti   美的\\midea
new_file_negative.save('D:\\#王琦\\方太烟灶-文本分析\\华帝\\vatti_com_negative.xlsx')
new_file_nutural.save('D:\\#王琦\\方太烟灶-文本分析\\华帝\\vatti_com_nutural.xlsx')

new_file_positive.close()
new_file_negative.close()
new_file_nutural.close()